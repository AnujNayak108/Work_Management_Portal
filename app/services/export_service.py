"""
Export Service — CSV/Excel generation for reports.
"""
import csv
import io
from openpyxl import Workbook
from ..models.goal import GoalSheet, Goal, Achievement
from ..models.user import User


def generate_achievement_csv(cycle_id):
    """Generate CSV of Planned vs Actual for all employees."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Employee ID', 'Employee Name', 'Department', 'Goal Title',
        'Thrust Area', 'Unit Type', 'Weightage %', 'Target Value',
        'Q1 Actual', 'Q1 Score', 'Q2 Actual', 'Q2 Score',
        'Q3 Actual', 'Q3 Score', 'Q4 Actual', 'Q4 Score', 'Status'
    ])

    sheets = GoalSheet.query.filter_by(cycle_id=cycle_id).all()
    for sheet in sheets:
        emp = sheet.employee
        for goal in sheet.goals:
            row = [
                emp.employee_id, emp.name,
                emp.department.name if emp.department else '',
                goal.title, goal.thrust_area, goal.unit_type,
                goal.weightage, goal.target_value or ''
            ]
            for q in ['Q1', 'Q2', 'Q3', 'Q4']:
                ach = next((a for a in goal.achievements if a.quarter == q), None)
                row.append(ach.actual_value if ach else '')
                row.append(round(ach.computed_score, 1) if ach and ach.computed_score else '')
            row.append(goal.status)
            writer.writerow(row)

    output.seek(0)
    return output.getvalue()


def generate_achievement_excel(cycle_id):
    """Generate Excel workbook of Planned vs Actual."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Achievement Report'

    headers = [
        'Employee ID', 'Employee Name', 'Department', 'Goal Title',
        'Thrust Area', 'Unit Type', 'Weightage %', 'Target Value',
        'Q1 Actual', 'Q1 Score', 'Q2 Actual', 'Q2 Score',
        'Q3 Actual', 'Q3 Score', 'Q4 Actual', 'Q4 Score', 'Status'
    ]
    ws.append(headers)

    # Style header row
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')

    sheets = GoalSheet.query.filter_by(cycle_id=cycle_id).all()
    for sheet in sheets:
        emp = sheet.employee
        for goal in sheet.goals:
            row = [
                emp.employee_id, emp.name,
                emp.department.name if emp.department else '',
                goal.title, goal.thrust_area, goal.unit_type,
                goal.weightage, goal.target_value or ''
            ]
            for q in ['Q1', 'Q2', 'Q3', 'Q4']:
                ach = next((a for a in goal.achievements if a.quarter == q), None)
                row.append(ach.actual_value if ach else '')
                row.append(round(ach.computed_score, 1) if ach and ach.computed_score else '')
            row.append(goal.status)
            ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
